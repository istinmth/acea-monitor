#!/usr/bin/env python3
import os
import re
import logging
import openpyxl
from openpyxl.styles import Font, Alignment
from copy import copy
from openpyxl.utils import get_column_letter, column_index_from_string

# Set up logging
logger = logging.getLogger('excel_formatter')

def format_excel_numbers(excel_path):
    """
    Post-process Excel file from Adobe API to ensure proper number formatting:
    - "," as thousands separator
    - "." as decimal separator
    
    Returns:
        bool: True if formatting was successful, False otherwise
    """
    try:
        logger.info(f"Post-processing Excel file: {excel_path}")
        
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_path)
        
        # Regex pattern to identify numbers
        number_pattern = re.compile(r'^-?\d{1,3}(,\d{3})*(\.\d+)?$|^-?\d+(\.\d+)?$')
        
        # Process each worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Process each cell
            for row in worksheet.iter_rows():
                for cell in row:
                    # Skip empty cells
                    if not cell.value:
                        continue
                    
                    # Handle string cells that might contain numbers
                    if isinstance(cell.value, str):
                        str_value = cell.value.strip()
                        
                        if number_pattern.match(str_value):
                            cleaned = str_value.replace(',', '')
                            
                            try:
                                if '.' in cleaned:
                                    cell.value = float(cleaned)
                                    cell.number_format = '#,##0.00'
                                else:
                                    cell.value = int(cleaned)
                                    cell.number_format = '#,##0'
                            except ValueError:
                                pass
                    
                    # Format existing numeric cells
                    elif isinstance(cell.value, int):
                        cell.number_format = '#,##0'
                    elif isinstance(cell.value, float):
                        cell.number_format = '#,##0.00'
        
        # Save the workbook
        workbook.save(excel_path)
        logger.info(f"Successfully formatted Excel file: {excel_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error formatting Excel file: {e}")
        return False

def extract_monthly_table(excel_path):
    """Extract the MONTHLY section to a new worksheet in the same Excel file."""
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        # Create a new worksheet
        if "Monthly" in wb.sheetnames:
            wb.remove(wb["Monthly"])
        monthly_ws = wb.create_sheet("Monthly")
        
        # Find the monthly table
        monthly_found = False
        
        for sheet_name in wb.sheetnames:
            if sheet_name == "Monthly":
                continue
                
            ws = wb[sheet_name]
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_values = [str(cell).upper() if cell is not None else "" for cell in row]
                row_text = " ".join(row_values)
                
                if "MONTHLY" in row_text:
                    monthly_start_row = row_idx
                    source_sheet = ws
                    monthly_found = True
                    break
            
            if monthly_found:
                break
        
        if monthly_found:
            # Find the end (YEAR TO DATE or blank row)
            monthly_end_row = None
            for row_idx in range(monthly_start_row, source_sheet.max_row + 1):
                row_values = [str(cell.value).upper() if cell.value is not None else "" 
                             for cell in source_sheet[row_idx]]
                row_text = " ".join(row_values)
                
                if all(cell.value is None or str(cell.value).strip() == "" 
                      for cell in source_sheet[row_idx]) or "YEAR TO DATE" in row_text:
                    monthly_end_row = row_idx - 1
                    break
            
            if not monthly_end_row:
                monthly_end_row = source_sheet.max_row
            
            # Copy the data
            for i, row_num in enumerate(range(monthly_start_row, monthly_end_row + 1), 1):
                for j, cell in enumerate(source_sheet[row_num], 1):
                    monthly_ws.cell(row=i, column=j).value = cell.value
                    monthly_ws.cell(row=i, column=j).number_format = cell.number_format
            
            # Auto-adjust column widths
            for column in monthly_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                monthly_ws.column_dimensions[column_letter].width = max_length + 4
            
            wb.save(excel_path)
            
            # After extracting, clean the table
            clean_monthly_table(excel_path)
            
            return True
        
        return False
        
    except Exception as e:
        logger.error(f"Error extracting monthly table: {e}")
        return False

def clean_monthly_table(excel_path, sheet_name="Monthly"):
    """
    Clean up the Monthly Excel table according to specific requirements:
    1. Delete all empty columns
    2. Empty specific columns (C, D, F, G, I, J, L, M, O, P, R, S, U, V)
    3. Delete any resulting empty columns
    4. Empty specific rows except for column A where the value is "EUROPEAN UNION", "EFTA", or "EU + EFTA + UK"
    
    Args:
        excel_path (str): Path to the Excel file
        sheet_name (str): Name of the worksheet to clean
    
    Returns:
        bool: True if cleaning was successful, False otherwise
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path)
        
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found")
            return False
            
        ws = wb[sheet_name]
        
        # Step 1: Delete all empty columns
        delete_empty_columns(ws)
        
        # Step 2: Empty specific columns (C, D, F, G, I, J, L, M, O, P, R, S, U, V)
        columns_to_empty = ['C', 'D', 'F', 'G', 'I', 'J', 'L', 'M', 'O', 'P', 'R', 'S', 'U', 'V']
        empty_specific_columns(ws, columns_to_empty)
        
        # Step 3: Delete any empty columns resulting from step 2
        delete_empty_columns(ws)
        
        # Step 4: Empty specific rows except column A
        special_rows = find_special_rows(ws, ["EUROPEAN UNION", "EFTA", "EU + EFTA + UK"])
        empty_rows_except_column_a(ws, special_rows)
        
        # Save the workbook
        wb.save(excel_path)
        logger.info(f"Successfully cleaned Excel file: {excel_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error cleaning Excel file: {e}")
        return False

def delete_empty_columns(worksheet):
    """Delete all empty columns in the worksheet."""
    # Identify empty columns
    empty_cols = []
    for col_idx in range(1, worksheet.max_column + 1):
        is_empty = True
        for row_idx in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None and str(cell_value).strip() != "":
                is_empty = False
                break
        if is_empty:
            empty_cols.append(col_idx)
    
    # Delete empty columns from right to left
    for col_idx in sorted(empty_cols, reverse=True):
        col_letter = get_column_letter(col_idx)
        worksheet.delete_cols(col_idx, 1)

def empty_specific_columns(worksheet, column_letters):
    """Empty the content of specific columns."""
    for col_letter in column_letters:
        try:
            col_idx = column_index_from_string(col_letter)
            # Check if column exists (may have shifted after deletions)
            if col_idx <= worksheet.max_column:
                for row_idx in range(1, worksheet.max_row + 1):
                    worksheet.cell(row=row_idx, column=col_idx).value = None
        except:
            # Column might not exist anymore due to previous deletions
            pass

def find_special_rows(worksheet, target_values):
    """Find rows where column A contains any of the target values."""
    special_rows = []
    for row_idx in range(1, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row_idx, column=1).value
        if cell_value is not None and any(target in str(cell_value).upper() for target in [t.upper() for t in target_values]):
            special_rows.append(row_idx)
    return special_rows

def empty_rows_except_column_a(worksheet, row_indices):
    """Empty all cells in specified rows except for column A."""
    for row_idx in row_indices:
        for col_idx in range(2, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).value = None