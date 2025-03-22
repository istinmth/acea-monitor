#!/usr/bin/env python3

import os
import sqlite3
import datetime
import logging
import sys
import json
from flask import Flask, render_template, send_from_directory, jsonify, request, Response, redirect, url_for
from apscheduler.schedulers.background import BackgroundScheduler
import scraper
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import re
import tempfile
import locale

# Set up logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler('/app/logs/app.log'), logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger('acea_webapp')

# Constants
DB_PATH = '/app/data/database.db'
PDF_DIR = '/app/data/pdfs'
EXCEL_DIR = '/app/data/excel'
LOG_FILE = '/app/logs/scraper.log'
CONVERSION_LOG = '/app/logs/conversion.log'

# Ensure directories exist
os.makedirs(PDF_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)

# Get the absolute path to the templates directory
template_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'templates'))
logger.info(f"Template directory path: {template_dir}")

# List files in the templates directory to verify
try:
    template_files = os.listdir(template_dir)
    logger.info(f"Template files found: {template_files}")
except Exception as e:
    logger.error(f"Error listing template files: {e}")
    template_files = []

# Create Flask app with explicit template folder
app = Flask(__name__, template_folder=template_dir)

def get_db_connection():
    """Create a database connection."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def run_scraper():
    """Run the scraper from the scheduler."""
    logger.info("Running scheduled scraper job")
    try:
        scraper.main()
    except Exception as e:
        logger.error(f"Error in scheduled scraper job: {e}")

# Initialize scheduler
scheduler = BackgroundScheduler()
scheduler.add_job(run_scraper, 'interval', hours=12)
scheduler.start()

def convert_pdf_to_excel(pdf_path, excel_path):
    """
    Convert a PDF to Excel with specific formatting:
    - Each page becomes a worksheet
    - "," as thousands separator
    - "." as decimal separator
    """
    try:
        # Set locale for number formatting
        locale.setlocale(locale.LC_NUMERIC, 'en_US.UTF-8')
        
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        
        # Remove the default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Extract base filename without extension for sheet naming
        base_filename = os.path.basename(pdf_path).replace('.pdf', '')
        
        # Open the PDF
        with pdfplumber.open(pdf_path) as pdf:
            # For each page in the PDF
            for i, page in enumerate(pdf.pages):
                # Create a worksheet for this page
                sheet_name = f"Page {i+1}"
                if i == 0:
                    sheet_name = "Summary"  # First page is usually a summary
                worksheet = workbook.create_sheet(title=sheet_name)
                
                # Extract tables from the page
                tables = page.extract_tables()
                
                # If no tables found, try to extract text and make a simple table
                if not tables:
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for row_idx, line in enumerate(lines, 1):
                            worksheet.cell(row=row_idx, column=1, value=line)
                    continue
                
                # Process each table
                row_offset = 1
                for table_idx, table in enumerate(tables):
                    if table_idx > 0:
                        row_offset += 2  # Add space between tables
                    
                    # Write each row of the table
                    for row_idx, row in enumerate(table, row_offset):
                        for col_idx, cell in enumerate(row, 1):
                            if cell is not None:
                                # Try to convert numbers and format them
                                cell_value = cell.strip() if isinstance(cell, str) else cell
                                
                                # Check if the cell appears to be a number
                                if isinstance(cell_value, str):
                                    # Remove any existing commas
                                    cell_value = cell_value.replace(',', '')
                                    
                                    # Try to convert to a number
                                    try:
                                        if '.' in cell_value:
                                            cell_value = float(cell_value)
                                        elif cell_value.isdigit():
                                            cell_value = int(cell_value)
                                    except ValueError:
                                        pass  # Keep as string if conversion fails
                                
                                # Write to the cell
                                cell_obj = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                                
                                # Format numbers with appropriate separators
                                if isinstance(cell_value, (int, float)):
                                    if isinstance(cell_value, int):
                                        cell_obj.number_format = '#,##0'  # Thousands separator
                                    else:
                                        cell_obj.number_format = '#,##0.00'  # Decimal and thousands
                                
                                # Style header cells (usually first row)
                                if row_idx == row_offset:
                                    cell_obj.font = Font(bold=True)
                                    cell_obj.alignment = Alignment(horizontal='center')
                    
                    row_offset = row_idx + 1
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    
                    # Set width with some padding
                    worksheet.column_dimensions[column_letter].width = max_length + 4
        
        # Save the workbook
        workbook.save(excel_path)
        return True
    except Exception as e:
        logger.error(f"Error converting PDF to Excel: {e}")
        with open(CONVERSION_LOG, 'a') as f:
            f.write(f"{datetime.datetime.now()}: Error converting {pdf_path}: {str(e)}\n")
        return False

def get_excel_path(pdf_path):
    """Get the Excel path for a PDF."""
    base_name = os.path.basename(pdf_path).replace('.pdf', '.xlsx')
    return os.path.join(EXCEL_DIR, base_name)

def ensure_excel_exists(report):
    """Ensure Excel file exists for a report, converting if needed."""
    pdf_path = report['pdf_path']
    excel_path = get_excel_path(pdf_path)
    
    # If Excel doesn't exist, create it
    if not os.path.exists(excel_path):
        logger.info(f"Converting PDF to Excel: {pdf_path}")
        convert_pdf_to_excel(pdf_path, excel_path)
    
    return excel_path

@app.route('/')
def index():
    """Render the homepage with latest reports and statistics."""
    conn = get_db_connection()
    
    # Get PC reports
    pc_reports = conn.execute(
        'SELECT * FROM reports WHERE type = "PC" ORDER BY publish_date DESC LIMIT 5'
    ).fetchall()
    
    # Get CV reports
    cv_reports = conn.execute(
        'SELECT * FROM reports WHERE type = "CV" ORDER BY publish_date DESC LIMIT 5'
    ).fetchall()
    
    # Get statistics
    stats = {
        'total_reports': conn.execute('SELECT COUNT(*) FROM reports').fetchone()[0],
        'pc_reports': conn.execute('SELECT COUNT(*) FROM reports WHERE type = "PC"').fetchone()[0],
        'cv_reports': conn.execute('SELECT COUNT(*) FROM reports WHERE type = "CV"').fetchone()[0]
    }
    
    # Get months with data
    months_with_data = conn.execute(
        '''SELECT strftime('%Y-%m', publish_date) as month, 
                  COUNT(*) as count 
           FROM reports 
           GROUP BY month 
           ORDER BY month DESC 
           LIMIT 12'''
    ).fetchall()
    
    conn.close()
    
    # Check when the last successful scan happened
    last_scan = "Unknown"
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r') as f:
                for line in reversed(f.readlines()):
                    if "Finished scanning for ACEA reports" in line:
                        last_scan = line.split(" - ")[0]
                        break
    except Exception as e:
        logger.error(f"Error reading log file: {e}")
    
    return render_template(
        'index.html', 
        pc_reports=pc_reports, 
        cv_reports=cv_reports,
        stats=stats,
        months=months_with_data,
        last_scan=last_scan
    )

@app.route('/reports/<report_type>')
def reports(report_type):
    """Show all reports of a specific type."""
    if report_type not in ['PC', 'CV']:
        return jsonify({'error': 'Invalid report type'}), 400
        
    conn = get_db_connection()
    reports = conn.execute(
        'SELECT * FROM reports WHERE type = ? ORDER BY publish_date DESC',
        (report_type,)
    ).fetchall()
    conn.close()
    
    return render_template(
        'reports.html', 
        reports=reports, 
        report_type=report_type
    )

@app.route('/pdf/<path:filename>')
def serve_pdf(filename):
    """Serve a PDF file."""
    return send_from_directory(PDF_DIR, filename)

@app.route('/excel/<path:filename>')
def serve_excel(filename):
    """Serve an Excel file."""
    return send_from_directory(EXCEL_DIR, filename)

@app.route('/convert/<int:report_id>')
def convert_report(report_id):
    """Convert a report's PDF to Excel and serve it."""
    conn = get_db_connection()
    report = conn.execute('SELECT * FROM reports WHERE id = ?', (report_id,)).fetchone()
    conn.close()
    
    if not report:
        return jsonify({'error': 'Report not found'}), 404
    
    pdf_path = report['pdf_path']
    excel_path = get_excel_path(pdf_path)
    
    # Convert PDF to Excel if not already done
    if not os.path.exists(excel_path):
        success = convert_pdf_to_excel(pdf_path, excel_path)
        if not success:
            return jsonify({'error': 'Failed to convert PDF to Excel'}), 500
    
    # Redirect to the Excel file
    return redirect(url_for('serve_excel', filename=os.path.basename(excel_path)))

@app.route('/api/stats')
def stats():
    """Return statistics about the reports."""
    conn = get_db_connection()
    
    pc_count = conn.execute('SELECT COUNT(*) FROM reports WHERE type = "PC"').fetchone()[0]
    cv_count = conn.execute('SELECT COUNT(*) FROM reports WHERE type = "CV"').fetchone()[0]
    
    latest_pc = conn.execute(
        'SELECT publish_date FROM reports WHERE type = "PC" ORDER BY publish_date DESC LIMIT 1'
    ).fetchone()
    
    latest_cv = conn.execute(
        'SELECT publish_date FROM reports WHERE type = "CV" ORDER BY publish_date DESC LIMIT 1'
    ).fetchone()
    
    conn.close()
    
    return jsonify({
        'total_reports': pc_count + cv_count,
        'pc_reports': pc_count,
        'cv_reports': cv_count,
        'latest_pc': latest_pc[0] if latest_pc else None,
        'latest_cv': latest_cv[0] if latest_cv else None
    })

@app.route('/run-scan', methods=['POST'])
def run_scan():
    """Manually trigger a scan."""
    try:
        scraper.main()
        return jsonify({'success': True, 'message': 'Scan completed successfully'})
    except Exception as e:
        logger.error(f"Error in manual scan: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/logs')
def view_logs():
    """View the application logs."""
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r') as f:
                logs = f.readlines()
                # Return only the last 100 lines
                logs = logs[-100:]
        else:
            logs = ["No logs found"]
    except Exception as e:
        logger.error(f"Error reading log file: {e}")
        logs = [f"Error reading logs: {e}"]
    
    return render_template('logs.html', logs=logs)

@app.route('/delete-reports', methods=['POST'])
def delete_reports():
    """Delete selected reports."""
    try:
        report_ids = request.json.get('report_ids', [])
        
        if not report_ids:
            return jsonify({'success': False, 'message': 'No reports selected'}), 400
        
        conn = get_db_connection()
        
        # Get file paths before deleting from database
        file_paths = []
        for report_id in report_ids:
            report = conn.execute('SELECT pdf_path FROM reports WHERE id = ?', (report_id,)).fetchone()
            if report and report['pdf_path']:
                file_paths.append(report['pdf_path'])
                
                # Also add Excel path if it exists
                excel_path = get_excel_path(report['pdf_path'])
                if os.path.exists(excel_path):
                    file_paths.append(excel_path)
        
        # Delete from database
        placeholders = ','.join(['?'] * len(report_ids))
        conn.execute(f'DELETE FROM reports WHERE id IN ({placeholders})', report_ids)
        conn.commit()
        conn.close()
        
        # Delete files
        for path in file_paths:
            try:
                if os.path.exists(path):
                    os.remove(path)
                    logger.info(f"Deleted file: {path}")
            except Exception as e:
                logger.error(f"Error deleting file {path}: {e}")
        
        return jsonify({
            'success': True, 
            'message': f'Successfully deleted {len(report_ids)} reports',
            'deleted_ids': report_ids
        })
    except Exception as e:
        logger.error(f"Error deleting reports: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/health')
def health_check():
    """Simple health check endpoint that doesn't require templates."""
    return jsonify({
        'status': 'ok',
        'timestamp': datetime.datetime.now().isoformat(),
        'template_dir': template_dir,
        'template_files': template_files,
        'db_exists': os.path.exists(DB_PATH),
        'pdf_dir_exists': os.path.exists(PDF_DIR)
    })

@app.route('/debug')
def debug_info():
    """Return debug information about the environment."""
    env_vars = {k: v for k, v in os.environ.items()}
    dirs = {
        'current': os.listdir('.'),
        'app': os.listdir('/app') if os.path.exists('/app') else [],
        'templates': template_files,
        'data': os.listdir('/app/data') if os.path.exists('/app/data') else []
    }
    
    return jsonify({
        'environment': env_vars,
        'directories': dirs,
        'python_path': sys.path
    })

@app.route('/convert-all', methods=['POST'])
def convert_all_pdfs():
    """Convert all PDFs to Excel format."""
    try:
        conn = get_db_connection()
        reports = conn.execute('SELECT * FROM reports').fetchall()
        conn.close()
        
        success_count = 0
        fail_count = 0
        
        for report in reports:
            pdf_path = report['pdf_path']
            excel_path = get_excel_path(pdf_path)
            
            # Skip if Excel already exists
            if os.path.exists(excel_path):
                success_count += 1
                continue
            
            # Convert PDF to Excel
            success = convert_pdf_to_excel(pdf_path, excel_path)
            if success:
                success_count += 1
            else:
                fail_count += 1
        
        return jsonify({
            'success': True,
            'message': f'Converted {success_count} PDFs to Excel, {fail_count} failed',
            'successes': success_count,
            'failures': fail_count
        })
    except Exception as e:
        logger.error(f"Error in batch conversion: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

if __name__ == '__main__':
    # Initialize database if it doesn't exist
    if not os.path.exists(DB_PATH):
        scraper.init_database()
    
    # Run the Flask app with debug enabled for troubleshooting
    app.run(host='0.0.0.0', port=9734, debug=True)